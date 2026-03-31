"""
Step 4: Put the scores back into the Excel file

Now that we've scored everything, we need to write the results back to 
an Excel file. We add two new columns - one for the score (0-4) and 
one for the explanation.

We make a backup of the original file first, just to be safe. Better
to have a copy than lose the original data!
"""

import json
import shutil
from datetime import datetime
from pathlib import Path
import pandas as pd
import openpyxl
import yaml


def populate_scores():
    """
    Main function - takes the scores from the JSON file and writes them 
    into the Excel workbook.
    """
    
    # Load our config to know where everythign is
    config_path = Path(__file__).parent.parent / "config" / "test2_config.yaml"
    with open(config_path, "r") as f:
        config = yaml.safe_load(f)
    
    # Figure out the file paths
    input_excel = Path(__file__).parent.parent.parent / "data" / "working" / "AutoEIT Sample Audio for Transcribing_WORKING.xlsx"
    scores_json = Path(__file__).parent.parent / "outputs" / "scores" / "all_scores.json"
    output_excel = Path(__file__).parent.parent / "data" / "output" / "scored_results.xlsx"
    
    print(f"Input Excel: {input_excel}")
    print(f"Scores JSON: {scores_json}")
    print(f"Output Excel: {output_excel}")
    
    # Make sure the files we're reading actually exist
    if not input_excel.exists():
        raise FileNotFoundError(f"Input Excel not found: {input_excel}")
    if not scores_json.exists():
        raise FileNotFoundError(f"Scores JSON not found: {scores_json}")
    
    # Create a backup of the original - this is good practice
    # Just in case something goes wrong, we have a copy
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = input_excel.parent / f"AutoEIT_Scored_{timestamp}.bak"
    shutil.copy2(input_excel, backup_path)
    print(f"Created backup: {backup_path}")
    
    # Load the scores from the JSON file we created in step 2
    with open(scores_json, "r", encoding="utf-8") as f:
        scores_data = json.load(f)
    
    print(f"Loaded {len(scores_data)} scores")
    
    # Group the scores by participant so we can write to the right sheet
    participants = {}
    for item in scores_data:
        p = item.get('participant')
        if p not in participants:
            participants[p] = []
        participants[p].append(item)
    
    # Open the Excel file
    wb = openpyxl.load_workbook(input_excel)
    
    # What should we call the new columns?
    score_col = config['SCORE_COLUMN']
    rationale_col = config['RATIONALE_COLUMN']
    
    print(f"\nWriting scores to sheets...")
    
    # Go through each participant and add their scores
    for participant, items in participants.items():
        if participant not in wb.sheetnames:
            print(f"  Warning: Sheet '{participant}' not found, skipping")
            continue
        
        ws = wb[participant]
        
        # Find where to put the new columns - we'll use D and E (after C)
        # Actually, let's figure out what column letters those are
        score_col_idx = 4  # Column D
        rationale_col_idx = 5  # Column E
        
        # Add the headers
        ws.cell(row=1, column=score_col_idx, value=score_col)
        ws.cell(row=1, column=rationale_col_idx, value=rationale_col)
        
        # Now write the actual scores
        scored_count = 0
        for item in items:
            item_num = item.get('item_number')
            score = item.get('score')
            rationale = item.get('rationale', '')
            
            # Find the right row - go through the sheet until we match the item number
            for row_idx in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=1).value
                if cell_val == item_num:
                    # Found the right row - write the score
                    ws.cell(row=row_idx, column=score_col_idx, value=score)
                    # And the explanation (shorten if it's really long)
                    if rationale:
                        rationale = rationale[:500] if len(rationale) > 500 else rationale
                        ws.cell(row=row_idx, column=rationale_col_idx, value=rationale)
                    scored_count += 1
                    break
        
        print(f"  {participant}: {scored_count}/30 scores written")
    
    # Save the modified workbook to the output location
    output_excel.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_excel)
    print(f"\n✅ Saved scored results to: {output_excel}")
    
    # Print a summary of what we just did
    print("\n" + "="*50)
    print("SCORE SUMMARY")
    print("="*50)
    
    scores = [item['score'] for item in scores_data if item.get('score') is not None]
    print(f"Total scored: {len(scores)}/{len(scores_data)}")
    
    if scores:
        print("\nScore distribution:")
        for s in range(5):
            count = scores.count(s)
            pct = count / len(scores) * 100
            print(f"  {s}: {count:3d} ({pct:5.1f}%)")
        
        print(f"\nMean score: {sum(scores)/len(scores):.2f}")
        print(f"Min: {min(scores)}, Max: {max(scores)}")


if __name__ == "__main__":
    populate_scores()